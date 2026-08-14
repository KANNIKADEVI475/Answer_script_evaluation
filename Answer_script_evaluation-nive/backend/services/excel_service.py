from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database.database import SessionLocal
from models.evaluation import Evaluation
from config.section_mapping import SECTION_MAPPING

import json


def _parse_marks(mark_string):
    if not isinstance(mark_string, str):
        return 0.0, 0.0

    parts = mark_string.split("/")
    if len(parts) != 2:
        try:
            return float(mark_string), 0.0
        except ValueError:
            return 0.0, 0.0

    obtained_str, maximum_str = parts
    try:
        obtained = float(obtained_str)
    except ValueError:
        obtained = 0.0

    try:
        maximum = float(maximum_str)
    except ValueError:
        maximum = 0.0

    return obtained, maximum


def export_excel():

    db = SessionLocal()

    evaluations = db.query(Evaluation).all()

    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Student Marks"

    # Header
    sheet.append([
        "S.No",
        "Register No",
        "Student Name",
        "Part A",
        "Part B",
        "Part C",
        "Part D",
        "Total"
    ])

    # Style Header
    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Fill student rows
    for index, evaluation in enumerate(evaluations, start=1):

        results = json.loads(evaluation.evaluation_data)

        section_marks = {
            "Part A": 0,
            "Part B": 0,
            "Part C": 0,
            "Part D": 0,
        }

        for i, item in enumerate(results):
            obtained, _ = _parse_marks(item.get("marks", "0/0"))

            for section, question_range in SECTION_MAPPING.items():

                if i in question_range:

                    section_marks[section] += obtained
                    break

        sheet.append([
            index,
            evaluation.register_number,
            evaluation.student_name,
            section_marks["Part A"],
            section_marks["Part B"],
            section_marks["Part C"],
            section_marks["Part D"],
            evaluation.total_marks
        ])

    # Auto-adjust column widths
    for column_cells in sheet.columns:

        length = max(len(str(cell.value or "")) for cell in column_cells)

        sheet.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = length + 5

    stream = BytesIO()

    workbook.save(stream)

    stream.seek(0)

    db.close()

    return stream